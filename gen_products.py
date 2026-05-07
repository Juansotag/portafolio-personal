import openpyxl, io

wb = openpyxl.load_workbook('Portafolio_Productos_GovLab.xlsx')
ws = wb['Portafolio GovLab']
headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
print("Headers:", headers)

def clean(v):
    if v is None:
        return 'NA'
    s = str(v).strip().replace('\n', ' ')
    if not s or s == 'None':
        return 'NA'
    return s

def js_str(v):
    # Escape single quotes and backslashes for JS single-quoted string
    return v.replace('\\', '\\\\').replace("'", "\\'")

products = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(v for v in row):
        continue
    p = {}
    for i, h in enumerate(headers):
        p[h] = clean(row[i])
    products.append(p)

print(f"Total products: {len(products)}")

out = io.open('products_generated.js', 'w', encoding='utf-8')
out.write('const PRODUCTS = [\n')

for p in products:
    seg  = p.get('Segmento', 'NA').replace('\n', ' ')
    tipo = p.get('Tipo', 'NA').replace('\n', ' ').replace(' & ', ' & ')
    name = p.get('Nombre del producto', 'NA')
    lema = p.get('Lema', 'NA')
    
    # Description key may have accent
    desc = 'NA'
    for k in p:
        if 'escripci' in k.lower():
            desc = p[k]
            break
    
    estado  = p.get('Estado', 'NA')
    ppt_raw = p.get('Link ppt', 'NA').strip()
    # Clean: strip descriptive text after .pdf, prepend assets/ if needed
    import re as _re
    ppt = _re.sub(r'\.pdf.*$', '.pdf', ppt_raw) if ppt_raw and ppt_raw != 'NA' else ppt_raw
    if ppt and ppt != 'NA' and not ppt.startswith('assets/') and not ppt.startswith('http'):
        ppt = 'assets/' + ppt
    linkweb = p.get('Link web', 'NA').strip()
    linkapp = p.get('Link App', 'NA').strip()
    github  = p.get('Link Github', 'NA').strip()
    logo    = p.get('Logo', 'NA').strip()
    video   = p.get('Video', 'NA').strip()

    # appUrl: prefer direct app, then Tableau web link
    if linkapp and linkapp != 'NA':
        appUrl = linkapp
    elif linkweb and linkweb != 'NA':
        appUrl = linkweb
    else:
        appUrl = ''

    # Logo path - normalize backslashes and fix known mismatches
    FALLBACK = 'assets/Govlab.png'
    FILENAME_FIXES = {
        'carresponde.png': 'caresponde.png',
        'alocandidato.png': 'alocandidato.jpg',
    }
    if logo and logo != 'NA':
        # Normalize backslashes to forward slashes
        logo_norm = logo.replace('\\', '/')
        # If already starts with 'assets/', use as-is; otherwise prepend
        if logo_norm.startswith('assets/'):
            logoPath = logo_norm
        else:
            logoPath = 'assets/' + logo_norm
        # Fix known filename mismatches
        for wrong, right in FILENAME_FIXES.items():
            if logoPath.endswith(wrong):
                logoPath = logoPath[:-len(wrong)] + right
    else:
        logoPath = FALLBACK


    # demoMode: if has ppt link → 'ppt', else 'mostrar'
    if ppt and ppt != 'NA':
        demoMode = 'ppt'
    else:
        demoMode = 'mostrar'

    def s(v):
        if not v or v == 'NA':
            return ''
        return js_str(v)

    out.write('  {\n')
    out.write(f"    segment: '{js_str(seg)}',\n")
    out.write(f"    type: '{js_str(tipo)}',\n")
    out.write(f"    name: '{js_str(name)}',\n")
    out.write(f"    slogan: '{js_str(lema) if lema != 'NA' else ''}',\n")
    out.write(f"    description: '{js_str(desc) if desc != 'NA' else ''}',\n")
    out.write(f"    status: '{js_str(estado)}',\n")
    out.write(f"    demoMode: '{demoMode}',\n")
    out.write(f"    appUrl: '{s(appUrl)}',\n")
    out.write(f"    githubUrl: '{s(github)}',\n")
    out.write(f"    logo: '{logoPath}',\n")
    out.write(f"    videoUrl: '{s(video)}',\n")
    out.write(f"    pptUrl: '{s(ppt)}',\n")
    out.write('  },\n')

out.write('];\n')
out.close()
print('Generated products_generated.js successfully')
